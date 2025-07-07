import pandas as pd
import json
import os
from openpyxl import load_workbook


def extract_key_value_pairs(excel_path, max_rows=100):
    """
    尽量从结构错乱的表格中提取出“字段 -> 值”的结构。
    """
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    data = []

    for row in sheet.iter_rows(min_row=1, max_row=max_rows):
        row_values = [cell.value if cell.value is not None else "" for cell in row]

        # 简单启发式：假设一行中有一个字段名 + 一个对应值
        non_empty = [(i, v) for i, v in enumerate(row_values) if str(v).strip() != ""]

        if len(non_empty) == 2:
            # 检测可能的键值对结构
            key, value = non_empty[0][1], non_empty[1][1]
            data.append({"字段": str(key).strip(), "值": str(value).strip()})

        elif len(non_empty) == 1 and data:
            # 如果只有一个值，可能是上一个字段的延续内容
            last_entry = data[-1]
            last_entry["值"] += " " + str(non_empty[0][1]).strip()

    return data


def save_to_json(data, output_path):
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    input_excel = "气动开关阀数据表-1.xlsx"  # 替换为你的文件名
    output_json = "结构化_清洗输出.json"

    if not os.path.exists(input_excel):
        print(f"❌ 文件未找到: {input_excel}")
    else:
        print("📥 读取中...")
        cleaned_data = extract_key_value_pairs(input_excel)
        save_to_json(cleaned_data, output_json)
        print(f"✅ 已导出清洗后的数据到: {output_json}")
