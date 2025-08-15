from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import pdfplumber as pb
import pandas as pd
from base import Base
import os
import re


class Extractor(Base):
    def __init__(self, path):
        super().__init__()
        self.file = path

    @staticmethod
    def load_files(directory):
        """load files"""
        paths = []
        for file in os.walk(directory):
            for f in file[2]:
                path = os.path.join(file[0], f)
                if os.path.isfile(path) and os.path.splitext(path)[1] == '.pdf':
                    paths.append(path)
        return paths

    def _load_data(self):
        if self.file and os.path.splitext(self.file)[1] == '.pdf':
            pdf = pb.open(self.file)
            page = pdf.pages[-1]
            words = page.extract_words(x_tolerance=5)
            lines = page.lines
            # convert coordination
            for index, word in enumerate(words):
                words[index]['y0'] = word['top']
                words[index]['y1'] = word['bottom']
            for index, line in enumerate(lines):
                lines[index]['x1'] = line['x0'] + line['width']
                lines[index]['y0'] = line['top']
                lines[index]['y1'] = line['bottom']
            return {'words': words, 'lines': lines}
        else:
            self.printLog("file %s can't be opened." % self.file)
            return None

    @staticmethod
    def _fill_line(lines):
        hlines = [line for line in lines if line['width'] > 0]  # 筛选横线
        hlines = sorted(hlines, key=lambda h: h['width'], reverse=True)[:-2]  # 剔除较短的两根
        vlines = [line for line in lines if line['height'] > 0]  # 筛选竖线
        vlines = sorted(vlines, key=lambda v: v['y0'])  # 按照坐标排列
        # 查找边框顶点
        hx0 = hlines[0]['x0']  # 左侧
        hx1 = hlines[0]['x1']  # 右侧
        vy0 = vlines[0]['y0']  # 顶部
        vy1 = vlines[-1]['y1']  # 底部

        thline = {'x0': hx0, 'y0': vy0, 'x1': hx1, 'y1': vy0}  # 顶部横线
        bhline = {'x0': hx0, 'y0': vy1, 'x1': hx1, 'y1': vy1}  # 底部横线
        lvline = {'x0': hx0, 'y0': vy0, 'x1': hx0, 'y1': vy1}  # 左侧竖线
        rvline = {'x0': hx1, 'y0': vy0, 'x1': hx1, 'y1': vy1}  # 右侧竖线

        hlines.insert(0, thline)
        hlines.append(bhline)
        vlines.insert(0, lvline)
        vlines.append(rvline)
        return {'hlines': hlines, 'vlines': vlines}

    @staticmethod
    def _is_point_in_rect(point, rect):
        """判断点是否在矩形内"""
        px, py = point
        p1, p2, p3, p4 = rect
        if p1[0] <= px <= p2[0] and p1[1] <= py <= p3[1]:
            return True
        else:
            return False

    @staticmethod
    def _find_cross_points(hlines, vlines):
        points = []
        delta = 1
        for vline in vlines:
            vx0 = vline['x0']
            vy0 = vline['y0']
            vy1 = vline['y1']
            for hline in hlines:
                hx0 = hline['x0']
                hy0 = hline['y0']
                hx1 = hline['x1']
                if (hx0 - delta) <= vx0 <= (hx1 + delta) and (vy0 - delta) <= hy0 <= (vy1 + delta):
                    points.append((int(vx0), int(hy0)))
        return points

    @staticmethod
    def _find_rects(cross_points):
        # 构造矩阵
        X = sorted(set([int(p[0]) for p in cross_points]))
        Y = sorted(set([int(p[1]) for p in cross_points]))
        df = pd.DataFrame(index=Y, columns=X)
        for p in cross_points:
            x, y = int(p[0]), int(p[1])
            df.loc[y, x] = 1
        df = df.infer_objects()
        df = df.fillna(0)
        # 寻找矩形
        rects = []
        COLS = len(df.columns) - 1
        ROWS = len(df.index) - 1
        for row in range(ROWS):
            for col in range(COLS):
                p0 = df.iat[row, col]  # 主点：必能构造一个矩阵
                cnt = col + 1
                while cnt <= COLS:
                    p1 = df.iat[row, cnt]
                    p2 = df.iat[row + 1, col]
                    p3 = df.iat[row + 1, cnt]
                    if p0 and p1 and p2 and p3:
                        rects.append(((df.columns[col], df.index[row]), (df.columns[cnt], df.index[row]), (
                            df.columns[col], df.index[row + 1]), (df.columns[cnt], df.index[row + 1])))
                        break
                    else:
                        cnt += 1
        return rects

    def _put_words_into_rect(self, words, rects):
        # 将words按照坐标层级放入矩阵中
        groups = {}
        delta = 2
        for word in words:
            p = (int(word['x0']), int((word['y0'] + word['y1']) / 2))
            flag = False
            for r in rects:
                if self._is_point_in_rect(p, r):
                    flag = True
                    groups[('IN', r[0][1], r)] = groups.get(
                        ('IN', r[0][1], r), []) + [word]
                    break
            if not flag:
                y_range = [
                              p[1] + x for x in range(delta)] + [p[1] - x for x in range(delta)]
                out_ys = [k[1] for k in list(groups.keys()) if k[0] == 'OUT']
                flag = False
                for y in set(y_range):
                    if y in out_ys:
                        v = out_ys[out_ys.index(y)]
                        groups[('OUT', v)].append(word)
                        flag = True
                        break
                if not flag:
                    groups[('OUT', p[1])] = [word]
        return groups

    @staticmethod
    def _find_text_by_same_line(group, delta=1):
        words = {}
        group = sorted(group, key=lambda x: x['x0'])
        for w in group:
            bottom = int(w['bottom'])
            text = w['text']
            k1 = [bottom - i for i in range(delta)]
            k2 = [bottom + i for i in range(delta)]
            k = set(k1 + k2)
            flag = False
            for kk in k:
                if kk in words:
                    words[kk] = words.get(kk, '') + text
                    flag = True
                    break
            if not flag:
                words[bottom] = words.get(bottom, '') + text
        return words

    def _split_words_into_diff_line(self, groups):
        groups2 = {}
        for k, g in groups.items():
            words = self._find_text_by_same_line(g, 3)
            groups2[k] = words
        return groups2

    @staticmethod
    def _index_of_y(x, rects):
        for index, r in enumerate(rects):
            if x == r[2][0][0]:
                return index + 1 if index + 1 < len(rects) else None
        return None

    @staticmethod
    def _find_outer(words):
        df = pd.DataFrame()
        for pos, text in words.items():
            if re.search(r'发票$', text):  # 发票名称
                df.loc[0, '发票名称'] = text
            elif re.search(r'发票代码', text):  # 发票代码
                num = ''.join(re.findall(r'[0-9]+', text))
                df.loc[0, '发票代码'] = num
            elif re.search(r'发票号码', text):  # 发票号码
                num = ''.join(re.findall(r'[0-9]+', text))
                df.loc[0, '发票号码'] = num
            elif re.search(r'开票日期', text):  # 开票日期
                date = ''.join(re.findall(
                    r'[0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日', text))
                df.loc[0, '开票日期'] = date
            elif '机器编号' in text and '校验码' in text:  # 校验码
                text1 = re.search(r'校验码:\d+', text)[0]
                num = ''.join(re.findall(r'[0-9]+', text1))
                df.loc[0, '校验码'] = num
                text2 = re.search(r'机器编号:\d+', text)[0]
                num = ''.join(re.findall(r'[0-9]+', text2))
                df.loc[0, '机器编号'] = num
            elif '机器编号' in text:
                num = ''.join(re.findall(r'[0-9]+', text))
                df.loc[0, '机器编号'] = num
            elif '校验码' in text:
                num = ''.join(re.findall(r'[0-9]+', text))
                df.loc[0, '校验码'] = num
            elif re.search(r'收款人', text):
                items = re.split(r'收款人:|复核:|开票人:|销售方:', text)
                items = [item for item in items if re.sub(
                    r'\s+', '', item) != '']
                df.loc[0, '收款人'] = items[0] if items and len(items) > 0 else ''
                df.loc[0, '复核'] = items[1] if items and len(items) > 1 else ''
                df.loc[0, '开票人'] = items[2] if items and len(items) > 2 else ''
                df.loc[0, '销售方'] = items[3] if items and len(items) > 3 else ''
        return df

    @staticmethod
    def _find_and_sort_rect_in_same_line(y, groups):
        same_rects_k = [k for k, v in groups.items() if k[1] == y]
        return sorted(same_rects_k, key=lambda x: x[2][0][0])

    def _find_inner(self, k, words, groups, groups2, free_zone_flag=False):
        df = pd.DataFrame()
        sort_words = sorted(words.items(), key=lambda x: x[0])
        text = [word for k, word in sort_words]
        context = ''.join(text)
        if '购买方' in context or '销售方' in context:
            y = k[1]
            x = k[2][0][0]
            same_rects_k = self._find_and_sort_rect_in_same_line(y, groups)
            target_index = self._index_of_y(x, same_rects_k)
            target_k = same_rects_k[target_index]
            group_context = groups2[target_k]
            prefix = '购买方' if '购买方' in context else '销售方'
            for pos, text in group_context.items():
                if '名称' in text:
                    name = re.sub(r'名称:', '', text)
                    name = re.sub(r'名称：', '', name)
                    df.loc[0, prefix + '名称'] = name
                elif '纳税人识别号' in text:
                    match = re.search(r'[:：]\s*([A-Za-z0-9]+)', text)
                    if match:
                        tax_man_id = match.group(1)
                        df.loc[0, prefix + '纳税人识别号'] = tax_man_id
                    else:
                        df.loc[0, prefix + '纳税人识别号'] = "/"
                elif '地址、电话' in text:
                    addr = re.sub(r'地址、电话:', '', text)
                    df.loc[0, prefix + '地址电话'] = addr
                elif '开户行及账号' in text:
                    account = re.sub(r'开户行及账号:', '', text)
                    df.loc[0, prefix + '开户行及账号'] = account
        elif '密码区' in context:
            y = k[1]
            x = k[2][0][0]
            same_rects_k = self._find_and_sort_rect_in_same_line(y, groups)
            target_index = self._index_of_y(x, same_rects_k)
            target_k = same_rects_k[target_index]
            words = groups2[target_k]
            context = [v for k, v in words.items()]
            context = ''.join(context)
            df.loc[0, '密码区'] = context
        elif '价税合计' in context:
            y = k[1]
            x = k[2][0][0]
            same_rects_k = self._find_and_sort_rect_in_same_line(y, groups)
            target_index = self._index_of_y(x, same_rects_k)
            target_k = same_rects_k[target_index]
            group_words = groups2[target_k]
            group_context = ''.join([w for k, w in group_words.items()])
            match = re.search(r'¥(\d+\.\d{2})', group_context)
            if match:
                amount = float(match.group(1).replace(',', ''))
                df.loc[0, '含税金额'] = amount
            else:
                df.loc[0, '含税金额'] = 0.0
        elif '项目名称' in context:
            match = re.search(r'合计¥([\d,]+\.\d{2})', context)
            if match:
                amount = float(match.group(1).replace(',', ''))
                df.loc[0, '未税金额'] = amount
            else:
                df.loc[0, '未税金额'] = 0.0
        elif '备注' in context:
            y = k[1]
            x = k[2][0][0]
            same_rects_k = self._find_and_sort_rect_in_same_line(y, groups)
            target_index = self._index_of_y(x, same_rects_k)
            if target_index:
                target_k = same_rects_k[target_index]
                group_words = groups2[target_k]
                group_context = ''.join([w for k, w in group_words.items()])
                df.loc[0, '备注'] = group_context
            else:
                df.loc[0, '备注'] = ''
        else:
            if free_zone_flag:
                return df, free_zone_flag
            y = k[1]
            same_rects_k = self._find_and_sort_rect_in_same_line(y, groups)
            if len(same_rects_k) == 8:
                free_zone_flag = True
                for kk in same_rects_k:
                    words = groups2[kk]
                    words = sorted(words.items(), key=lambda x: x[0]) if words and len(
                        words) > 0 else None
                    key = words[0][1] if words and len(words) > 0 else None
                    val = [word[1] for word in words[1:]
                           ] if key and words and len(words) > 1 else ''
                    val = '\n'.join(val) if val else ''
                    if key:
                        df.loc[0, key] = val
        return df, free_zone_flag

    def extract(self):
        data = self._load_data()
        words = data['words']
        lines = data['lines']
        lines = self._fill_line(lines)
        hlines = lines['hlines']
        vlines = lines['vlines']

        cross_points = self._find_cross_points(hlines, vlines)
        rects = self._find_rects(cross_points)
        word_groups = self._put_words_into_rect(words, rects)
        word_groups2 = self._split_words_into_diff_line(word_groups)
        df = pd.DataFrame()
        free_zone_flag = False
        for k, words in word_groups2.items():
            if k[0] == 'OUT':
                df_item = self._find_outer(words)
            else:
                df_item, free_zone_flag = self._find_inner(
                    k, words, word_groups, word_groups2, free_zone_flag)
            df = pd.concat([df, df_item], axis=1)
        return df


if __name__ == '__main__':
    IN_PATH = '../input/InvoiceInformation/'
    OUT_PATH = '../output/20250425/发票下载情况统计.xlsx'
    files_path = Extractor('').load_files(IN_PATH)
    num = len(files_path)
    data = pd.DataFrame()
    for index, file_path in enumerate(files_path):
        print(f'{index + 1}/{num}({round((index + 1) / num * 100, 2)}%)\t{file_path}')
        extractor = Extractor(file_path)
        try:
            d = extractor.extract()
            data = pd.concat([data, d], axis=0, sort=False, ignore_index=True)
        except Exception as e:
            print('file error:', file_path, '\n', e)
    # columns_order = ["发票号码", "开票日期", "购买方名称", "购买方纳税人识别号", "销售方名称", "销售方纳税人识别号", "未税金额", "含税金额", "备注"]
    columns_order = ["发票号码", "购买方名称", "销售方名称", "未税金额", "含税金额"]
    data = data[columns_order]
    data.to_excel(OUT_PATH, sheet_name='发票下载情况统计', index=False)

    # 设置所有 sheet 格式
    wb = load_workbook(OUT_PATH)
    ws = wb.active

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    wb.save(OUT_PATH)
